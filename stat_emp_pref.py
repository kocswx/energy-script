# coding:utf-8
from openpyxl import Workbook
from config.config_db_server import *
from _openpyxl_constant import *
from _util import *
import datetime

sheet_title = '员工绩效报表'
cell_name = ['员工ID', '员工姓名', '职位', '卡号', '状态', '充值金额', '退款金额',
             '销售笔数', '销售升数', '销售金额', '推荐会员数', '所属站点']


class PrefInfo:
    charge_amount = 0
    refund_amount = 0
    sale_cnt = 0
    sale_vol = 0
    sale_amount = 0
    referrer_cnt = 0


def _get_emp_record(group_id, end_time, station_id):
    sql = "SELECT EMP_ID,EMP_NAME,TITLE,CARD_NO,STATUS,STATION_NAME from eng_erp.emp_info " \
          "where GROUP_ID=%d AND CREATED_TIME<'%s'" % (group_id, end_time)
    if station_id:
        sql = sql + 'AND STATION_ID=%d' % station_id
    sql += " order by EMP_ID"
    cur.execute(sql)
    return cur.fetchall()


def _get_emp_pref(group_id, emp_id, begin_time, end_time):
    pref = PrefInfo()
    sql = "SELECT IFNULL(SUM(AMT+GIFT_AMT),0) from eng_crm.history_charge where GROUP_ID=%d AND OPT_EMP_ID='%s' " \
          "and CREATED_TIME>='%s' and CREATED_TIME<'%s'" % (group_id, emp_id, begin_time, end_time)
    # print(sql)
    cur.execute(sql)
    pref.charge_amount = cur.fetchone()[0]

    sql = "SELECT IFNULL(SUM(Amt),0) FROM eng_crm.history_refund WHERE GROUP_ID=%d AND OPT_EMP_ID='%s' " \
          "and CREATED_TIME>='%s' and CREATED_TIME<'%s'" % (group_id, emp_id, begin_time, end_time)
    # print(sql)
    cur.execute(sql)
    pref.refund_amount = cur.fetchone()[0]

    sql = "SELECT count(1),IFNULL(SUM(VOL),0),IFNULL(SUM(RECE_AMT),0) FROM eng_order.fuel_order WHERE GROUP_ID=%d " \
          "AND EMP_ID='%s'AND CREATED_TIME>='%s' AND CREATED_TIME<'%s'" % (group_id, emp_id, begin_time, end_time)
    # print(sql)
    cur.execute(sql)
    sale = cur.fetchone()
    pref.sale_cnt = sale[0]
    pref.sale_vol = sale[1]
    pref.sale_amount = sale[2]

    sql = "SELECT COUNT(1) from eng_crm.member_info WHERE GROUP_ID=%d and REFERRER_TYPE='EMP' " \
          "AND REFERRER_ID='%s' AND CREATED_TIME>='%s' AND CREATED_TIME<'%s'" % (
              group_id, emp_id, begin_time, end_time)
    # print(sql)
    cur.execute(sql)
    pref.referrer_cnt = cur.fetchone()[0]

    return pref


def _write_excel(group_id, begin_time, end_time, station_id, emp_name, save_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
    row_index = 1

    sheet.merge_cells("A1:%s1" % column_letter[len(cell_name) - 1])
    set_cell(sheet, row_index, 1, sheet.title, font_rpt_title, align_center, border)
    row_index += 1

    set_cell(sheet, row_index, 1, ('统计时间：%s 到 %s' % (begin_time, end_time)), font_rpt_normal, align_left, border)
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
    set_cell(sheet, row_index, 5, ('提交员工：%s' % emp_name), font_rpt_normal, align_center, border)
    sheet.merge_cells(start_row=row_index, start_column=5, end_row=row_index, end_column=8)
    set_cell(sheet, row_index, 9, ('生成时间：%s' % get_now_full()), font_rpt_normal, align_right, border)
    sheet.merge_cells(start_row=row_index, start_column=9, end_row=row_index, end_column=12)
    row_index += 1

    for i in range(len(cell_name)):
        set_cell(sheet, row_index, i + 1, cell_name[i], font_rpt_cell_header, align_center, border)
        sheet.column_dimensions[column_letter[i]].width = 12
    row_index += 1

    emp_list = _get_emp_record(group_id, end_time, station_id)
    # print(member_list)
    for item in emp_list:
        emp_id = item[0]
        print(emp_id)
        pref = _get_emp_pref(group_id, emp_id, begin_time, end_time)
        set_cell(sheet, row_index, 1, emp_id, font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 2, item[1], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 3, item[2], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 4, item[3], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 5, MEMBER_STATUS_DICT[item[4]], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 6, pref.charge_amount, font_rpt_normal, align_center, border)  # 充值金额
        set_cell(sheet, row_index, 7, pref.refund_amount, font_rpt_normal, align_center, border)  # 退款金额
        set_cell(sheet, row_index, 8, pref.sale_cnt, font_rpt_normal, align_center, border)  # 销售笔数
        set_cell(sheet, row_index, 9, pref.sale_vol, font_rpt_normal, align_center, border)  # 销售升数
        set_cell(sheet, row_index, 10, pref.sale_amount, font_rpt_normal, align_center, border)  # 销售金额
        set_cell(sheet, row_index, 11, pref.referrer_cnt, font_rpt_normal, align_center, border)  # 推荐会员数
        set_cell(sheet, row_index, 12, item[5], font_rpt_normal, align_center, border)
        row_index += 1

    wb.save(save_path)


cur = conn.cursor()

if __name__ == '__main__':
    print("stat_emp_pref start: %s" % (datetime.datetime.now()))
    cur.execute("select ID,GROUP_ID,BEGIN_TIME,END_TIME,STATION_ID,EMP_NAME from eng_report.stat_job "
                "where REPORT_TYPE='empPerf' and CALC_REPORT=0 limit 1")
    job = cur.fetchone()
    if job:
        print(job)
        # /opt/eng_static/report/10003/empPref-25.xlsx
        now_year_month = get_now_year_month()
        save_path = "%s/%s/%s/empPref-%d.xlsx" % (save_report_folder, job[1], now_year_month, job[0])
        return_path = "%s/%s/%s/empPref-%d.xlsx" % (return_report_prefix, job[1], now_year_month, job[0])
        print(save_path)
        prepare_path(save_path)
        station_id = job[4]
        if station_id is None or station_id <= 0:
            station_id = None
        _write_excel(job[1], job[2], job[3], station_id, job[5], save_path)
        cur.execute("update eng_report.stat_job set CALC_REPORT=1,REPORT_PATH='%s' where ID=%d" % (return_path, job[0]))
        conn.commit()
    print("stat_emp_pref end: %s" % (datetime.datetime.now()))
    cur.close()
    conn.close()
