from openpyxl.styles import Font, Alignment, Border, Side

font_rpt_title = Font(name=u'宋体', bold=True, size=18)
font_rpt_cell_header = Font(name=u'宋体', bold=True)
font_rpt_normal = Font(name=u'宋体')

align_left = Alignment(horizontal='left', vertical='center')
align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

border = Border(left=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))

column_letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'G', 'K', 'L', 'M',
                 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AG', 'AK', 'AL', 'AM',
                 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

'''
列宽最小值为 8.38；而行高最小值为 15
# 设置行高
sheet['A1']='行高被设置为 100'
sheet.row_dimensions[1].height=100

# 设置列宽
sheet['B2']='列宽被设置为 50'
sheet.column_dimensions['B'].width=50
'''


def set_cell(sheet, row_index, cell_index, value, font_style, alignment_style, border_style):
    sheet.cell(row_index, cell_index).font = font_style
    sheet.cell(row_index, cell_index).alignment = alignment_style
    sheet.cell(row_index, cell_index).border = border_style
    sheet.cell(row_index, cell_index, value)


if __name__ == '__main__':
    print(column_letter)
    pass
