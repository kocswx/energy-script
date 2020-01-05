import datetime
from openpyxl import Workbook
from config_openpyxl import *
from util import *


class Total:
    cnt = 0
    vol = 0
    receAmt = 0
    realAmt = 0
    discAmt = 0

    chargeAmt = 0
    giftAmt = 0
    totalAmt = 0

    rowCnt = 0
    rowCnt = 0


# 油品-油枪销售汇总
def _shift_prod_noz_order(cur, ctx, shift_no, shift_date):
    print('rpt_shift_prod_noz_order:',
          cur.execute(
              "DELETE FROM rpt_shift_prod_noz_order WHERE STATION_ID='%s' AND SHIFT_NO='%s'" % (ctx[1], shift_no)))
    sql = "INSERT INTO `rpt_shift_prod_noz_order`(`GROUP_ID`, `STATION_ID`, `SHIFT_NO`, `SHIFT_DATE`, `NOZ_NO`, `PROD_ID`,`PROD_NAME`, `CNT`, `VOL`, `RECE_AMT`, `REAL_AMT`, `DISC_AMT`)" \
          "SELECT GROUP_ID,STATION_ID,SHIFT_NO,'%s' AS SHIFT_DATE,NOZ_NO,PROD_ID,PROD_NAME,COUNT(1) AS CNT,ifnull(SUM(VOL), 0) AS VOL,ifnull(SUM(RECE_AMT), 0) AS RECE_AMT," \
          "ifnull(SUM(REAL_AMT), 0) AS REAL_AMT,ifnull(SUM(DISC_AMT), 0) AS DISC_AMT FROM fuel_order WHERE STATION_ID='%s' AND SHIFT_NO ='%s' GROUP BY GROUP_ID,STATION_ID,SHIFT_NO,NOZ_NO,PROD_ID,PROD_NAME" % (
              shift_date, ctx[1], shift_no)
    print(sql)
    cur.execute(sql)


# 员工-油品-支付方式汇总 非余额支付
def _shift_emp_prod_pay(cur, ctx, shift_no, shift_date):
    print('shift_emp_prod_pay:',
          cur.execute(
              "DELETE FROM rpt_shift_emp_prod_pay WHERE STATION_ID='%s' AND SHIFT_NO='%s'" % (ctx[1], shift_no)))
    sql = "INSERT INTO `rpt_shift_emp_prod_pay` (`GROUP_ID`, `STATION_ID`, `SHIFT_NO`, `SHIFT_DATE`, `PROD_ID`, `PROD_NAME`, `EMP_ID`, `EMP_NAME`, `PAY_TYPE_ID`, `PAY_TYPE_NAME`, `CNT`, `VOL`, `RECE_AMT`, `REAL_AMT`, `DISC_AMT`) " \
          "SELECT GROUP_ID,STATION_ID,SHIFT_NO,'%s' as SHIFT_DATE,PROD_ID,PROD_NAME,EMP_ID,EMP_NAME,PAY_TYPE_ID,PAY_TYPE_NAME,COUNT(1) AS CNT,SUM(VOL) AS VOL,SUM(RECE_AMT) AS RECE_AMT, SUM(REAL_AMT) AS REAL_AMT,SUM(DISC_AMT) AS DISC_AMT " \
          "FROM fuel_order WHERE  STATION_ID='%s' AND PAY_TYPE_ID<>2 and SHIFT_NO='%s' GROUP BY EMP_ID,EMP_NAME,PROD_ID,PROD_NAME,PAY_TYPE_ID,PAY_TYPE_NAME, GROUP_ID,STATION_ID,SHIFT_NO" % (
              shift_date, ctx[1], shift_no)
    print(sql)
    cur.execute(sql)


def _create_excel(cur, ctx, shift_no):
    shift_sql = "select SHIFT_NO,SHIFT_DATE,START_TIME,END_TIME,EMP_NAME FROM shift_record " \
                "where GROUP_ID='%s' AND STATION_ID='%s' AND SHIFT_NO='%s'" % (ctx[0], ctx[1], shift_no)
    cur.execute(shift_sql)
    shift_row = cur.fetchone()
    print(shift_row)
    wb = Workbook()
    sheet = wb.active
    sheet.title = '油品-油枪销售汇总'
    cell_name = ['油品名称', '枪号', '笔数', '销售升数', '销售金额', '实收金额', '优惠金额']
    row_index = 1

    sheet.merge_cells("A1:%s1" % column_letter[len(cell_name) - 1])
    sheet.cell(row_index, 1).value = sheet.title
    sheet.cell(row_index, 1).font = font_rpt_title
    sheet.cell(row_index, 1).alignment = align_center
    sheet.cell(row_index, 1).border = border
    row_index += 1
    sheet.cell(row_index, 1, '站点名称：%s' % ctx[2])
    sheet.cell(row_index, 1).alignment = align_left
    sheet.cell(row_index, 1).border = border
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
    sheet.cell(row_index, 5, '交班员工：%s' % shift_row[4])
    sheet.cell(row_index, 5).alignment = align_right
    sheet.cell(row_index, 5).border = border
    sheet.merge_cells(start_row=row_index, start_column=5, end_row=row_index, end_column=len(cell_name))
    row_index += 1

    sheet.cell(row_index, 1, '班次号：%s' % shift_row[0])
    sheet.cell(row_index, 1).alignment = align_left
    sheet.cell(row_index, 1).border = border
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
    sheet.cell(row_index, 3, '班次日期：%s' % shift_row[1])
    sheet.cell(row_index, 3).alignment = align_center
    sheet.cell(row_index, 3).border = border
    sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=4)
    sheet.cell(row_index, 5, '打印时间：%s' % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    sheet.cell(row_index, 5).alignment = align_right
    sheet.cell(row_index, 5).border = border
    sheet.merge_cells(start_row=row_index, start_column=5, end_row=row_index, end_column=7)
    row_index += 1

    for i in range(len(cell_name)):
        sheet.cell(row_index, i + 1).font = font_rpt_cell_header
        sheet.cell(row_index, i + 1).alignment = align_center
        sheet.cell(row_index, i + 1).border = border
        sheet.cell(row_index, i + 1, cell_name[i])
        sheet.column_dimensions[column_letter[i]].width = 12
    row_index += 1

    total = Total()  # 总计
    prod_total = Total()  # 行汇总

    rpt_sql = "SELECT PROD_NAME,NOZ_NO,CNT,VOL,RECE_AMT,REAL_AMT,DISC_AMT,PROD_ID FROM rpt_shift_prod_noz_order " \
              "WHERE STATION_ID='%s' and SHIFT_NO='%s' ORDER BY PROD_ID,NOZ_NO" % (ctx[1], shift_no)
    print(rpt_sql)
    cur.execute(rpt_sql)
    rows = cur.fetchall()
    if rows:
        for data_index in range(len(rows)):
            row = rows[data_index]
            prod = row[0]
            for i in range(len(cell_name)):
                sheet.cell(row_index, i + 1).font = font_rpt_normal
                sheet.cell(row_index, i + 1).alignment = align_center
                sheet.cell(row_index, i + 1).border = border
                sheet.cell(row_index, i + 1, row[i])
            row_index += 1
            # 总累计
            total.cnt += row[2]
            total.vol += row[3]
            total.receAmt += row[4]
            total.realAmt += row[5]
            total.discAmt += row[6]
            # 行累计
            prod_total.cnt += row[2]
            prod_total.vol += row[3]
            prod_total.receAmt += row[4]
            prod_total.realAmt += row[5]
            prod_total.discAmt += row[6]
            prod_total.rowCnt += 1
            if data_index > 0 and (data_index + 1 == len(rows) or prod != rows[data_index + 1][0]):
                if prod_total.rowCnt > 1:
                    merge_start_row = row_index - prod_total.rowCnt
                    sheet.merge_cells(start_row=merge_start_row, start_column=1, end_row=row_index - 1, end_column=1)
                # 写油品小计
                sheet.cell(row_index, 1).font = font_rpt_normal
                sheet.cell(row_index, 1).alignment = align_center
                sheet.cell(row_index, 1).border = border
                sheet.cell(row_index, 1, '小计：')
                sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
                sheet.cell(row_index, 3).font = font_rpt_normal
                sheet.cell(row_index, 3).alignment = align_center
                sheet.cell(row_index, 3).border = border
                sheet.cell(row_index, 3, prod_total.cnt)
                sheet.cell(row_index, 4).font = font_rpt_normal
                sheet.cell(row_index, 4).alignment = align_center
                sheet.cell(row_index, 4).border = border
                sheet.cell(row_index, 4, prod_total.vol)
                sheet.cell(row_index, 5).font = font_rpt_normal
                sheet.cell(row_index, 5).alignment = align_center
                sheet.cell(row_index, 5).border = border
                sheet.cell(row_index, 5, prod_total.receAmt)
                sheet.cell(row_index, 6).font = font_rpt_normal
                sheet.cell(row_index, 6).alignment = align_center
                sheet.cell(row_index, 6).border = border
                sheet.cell(row_index, 6, prod_total.realAmt)
                sheet.cell(row_index, 7).font = font_rpt_normal
                sheet.cell(row_index, 7).alignment = align_center
                sheet.cell(row_index, 7).border = border
                sheet.cell(row_index, 7, prod_total.discAmt)
                prod_total = Total()
                row_index += 1
    # 总计
    sheet.cell(row_index, 1).font = font_rpt_normal
    sheet.cell(row_index, 1).alignment = align_center
    sheet.cell(row_index, 1).border = border
    sheet.cell(row_index, 1, '总计：')
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
    sheet.cell(row_index, 3).font = font_rpt_normal
    sheet.cell(row_index, 3).alignment = align_center
    sheet.cell(row_index, 3).border = border
    sheet.cell(row_index, 3, total.cnt)
    sheet.cell(row_index, 4).font = font_rpt_normal
    sheet.cell(row_index, 4).alignment = align_center
    sheet.cell(row_index, 4).border = border
    sheet.cell(row_index, 4, total.vol)
    sheet.cell(row_index, 5).font = font_rpt_normal
    sheet.cell(row_index, 5).alignment = align_center
    sheet.cell(row_index, 5).border = border
    sheet.cell(row_index, 5, total.receAmt)
    sheet.cell(row_index, 6).font = font_rpt_normal
    sheet.cell(row_index, 6).alignment = align_center
    sheet.cell(row_index, 6).border = border
    sheet.cell(row_index, 6, total.realAmt)
    sheet.cell(row_index, 7).font = font_rpt_normal
    sheet.cell(row_index, 7).alignment = align_center
    sheet.cell(row_index, 7).border = border
    sheet.cell(row_index, 7, total.discAmt)
    row_index += 1
    # 统一行高
    for row_index in range(row_index):
        sheet.row_dimensions[row_index].height = 20
    sheet.protection.sheet = True
    sheet.protection.password = '123456'
    sheet.protection.enable()

    sheet = wb.create_sheet()
    sheet.title = '员工加油-支付方式汇总'
    cell_name = ['加油员', '支付方式', '笔数', '销售升数', '销售金额', '实收金额', '优惠金额']
    row_index = 1
    sheet.merge_cells("A1:%s1" % column_letter[len(cell_name) - 1])
    sheet.cell(row_index, 1).value = sheet.title
    sheet.cell(row_index, 1).font = font_rpt_title
    sheet.cell(row_index, 1).alignment = align_center
    sheet.cell(row_index, 1).border = border
    row_index += 1

    sheet.cell(row_index, 1, '站点名称：%s' % ctx[2])
    sheet.cell(row_index, 1).alignment = align_left
    sheet.cell(row_index, 1).border = border
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=4)
    sheet.cell(row_index, 5, '交班员工：%s' % shift_row[4])
    sheet.cell(row_index, 5).alignment = align_right
    sheet.cell(row_index, 5).border = border
    sheet.merge_cells(start_row=row_index, start_column=5, end_row=row_index, end_column=len(cell_name))
    row_index += 1

    sheet.cell(row_index, 1, '班次号：%s' % shift_row[0])
    sheet.cell(row_index, 1).alignment = align_left
    sheet.cell(row_index, 1).border = border
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
    sheet.cell(row_index, 3, '班次日期：%s' % shift_row[1])
    sheet.cell(row_index, 3).alignment = align_center
    sheet.cell(row_index, 3).border = border
    sheet.merge_cells(start_row=row_index, start_column=3, end_row=row_index, end_column=4)
    sheet.cell(row_index, 5, '打印时间：%s' % datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    sheet.cell(row_index, 5).alignment = align_right
    sheet.cell(row_index, 5).border = border
    sheet.merge_cells(start_row=row_index, start_column=5, end_row=row_index, end_column=7)
    row_index += 1

    for i in range(len(cell_name)):
        sheet.cell(row_index, i + 1).font = font_rpt_cell_header
        sheet.cell(row_index, i + 1).alignment = align_center
        sheet.cell(row_index, i + 1).border = border
        sheet.cell(row_index, i + 1, cell_name[i])
        sheet.column_dimensions[column_letter[i]].width = 12
    row_index += 1

    total = Total()  # 总计
    prod_total = Total()  # 行汇总
    rpt_sql = "SELECT EMP_NAME,PAY_TYPE_NAME,SUM(CNT) AS CNT,SUM(VOL)AS VOL,SUM(RECE_AMT)AS RECE_AMT,SUM(REAL_AMT)AS REAL_AMT," \
              "SUM(DISC_AMT)AS DISC_AMT FROM rpt_shift_emp_prod_pay WHERE STATION_ID='%s' and SHIFT_NO='%s'" \
              "GROUP BY EMP_NAME,PAY_TYPE_NAME" % (ctx[1], shift_no)
    print(rpt_sql)
    cur.execute(rpt_sql)
    rows = cur.fetchall()
    if rows:
        for data_index in range(len(rows)):
            row = rows[data_index]
            prod = row[0]
            for i in range(len(cell_name)):
                sheet.cell(row_index, i + 1).font = font_rpt_normal
                sheet.cell(row_index, i + 1).alignment = align_center
                sheet.cell(row_index, i + 1).border = border
                sheet.cell(row_index, i + 1, row[i])
            row_index += 1
            # 总累计
            total.cnt += row[2]
            total.vol += row[3]
            total.receAmt += row[4]
            total.realAmt += row[5]
            # 行累计
            prod_total.cnt += row[2]
            prod_total.vol += row[3]
            prod_total.receAmt += row[4]
            prod_total.realAmt += row[5]
            prod_total.rowCnt += 1
            if data_index > 0 and (data_index + 1 == len(rows) or prod != rows[data_index + 1][0]):
                if prod_total.rowCnt > 1:
                    merge_start_row = row_index - prod_total.rowCnt
                    sheet.merge_cells(start_row=merge_start_row, start_column=1, end_row=row_index - 1, end_column=1)
                # 写油品小计
                sheet.cell(row_index, 1).font = font_rpt_normal
                sheet.cell(row_index, 1).alignment = align_center
                sheet.cell(row_index, 1).border = border
                sheet.cell(row_index, 1, '小计：')
                sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
                sheet.cell(row_index, 3).font = font_rpt_normal
                sheet.cell(row_index, 3).alignment = align_center
                sheet.cell(row_index, 3).border = border
                sheet.cell(row_index, 3, prod_total.cnt)
                sheet.cell(row_index, 4).font = font_rpt_normal
                sheet.cell(row_index, 4).alignment = align_center
                sheet.cell(row_index, 4).border = border
                sheet.cell(row_index, 4, prod_total.vol)
                sheet.cell(row_index, 5).font = font_rpt_normal
                sheet.cell(row_index, 5).alignment = align_center
                sheet.cell(row_index, 5).border = border
                sheet.cell(row_index, 5, prod_total.receAmt)
                sheet.cell(row_index, 6).font = font_rpt_normal
                sheet.cell(row_index, 6).alignment = align_center
                sheet.cell(row_index, 6).border = border
                sheet.cell(row_index, 6, prod_total.realAmt)
                sheet.cell(row_index, 7).font = font_rpt_normal
                sheet.cell(row_index, 7).alignment = align_center
                sheet.cell(row_index, 7).border = border
                sheet.cell(row_index, 7, prod_total.discAmt)
                prod_total = Total()
                row_index += 1
    # 总计
    sheet.cell(row_index, 1).font = font_rpt_normal
    sheet.cell(row_index, 1).alignment = align_center
    sheet.cell(row_index, 1).border = border
    sheet.cell(row_index, 1, '总计：')
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=2)
    sheet.cell(row_index, 3).font = font_rpt_normal
    sheet.cell(row_index, 3).alignment = align_center
    sheet.cell(row_index, 3).border = border
    sheet.cell(row_index, 3, total.cnt)
    sheet.cell(row_index, 4).font = font_rpt_normal
    sheet.cell(row_index, 4).alignment = align_center
    sheet.cell(row_index, 4).border = border
    sheet.cell(row_index, 4, total.vol)
    sheet.cell(row_index, 5).font = font_rpt_normal
    sheet.cell(row_index, 5).alignment = align_center
    sheet.cell(row_index, 5).border = border
    sheet.cell(row_index, 5, total.receAmt)
    sheet.cell(row_index, 6).font = font_rpt_normal
    sheet.cell(row_index, 6).alignment = align_center
    sheet.cell(row_index, 6).border = border
    sheet.cell(row_index, 6, total.realAmt)
    sheet.cell(row_index, 7).font = font_rpt_normal
    sheet.cell(row_index, 7).alignment = align_center
    sheet.cell(row_index, 7).border = border
    sheet.cell(row_index, 7, total.discAmt)
    row_index += 1
    # 统一行高
    for row_index in range(row_index):
        sheet.row_dimensions[row_index].height = 20
    sheet.protection.sheet = True
    sheet.protection.password = '123456'
    sheet.protection.enable()

    wb.security.workbookPassword = '123456'
    wb.security.lockStructure = True
    return wb


def build_shift_report(conn, argv):
    print(argv)
    station_id = argv[1]
    shift_no = argv[2]
    shift_date = argv[3]
    cur = conn.cursor()
    fetch_station = "SELECT GROUP_ID,STATION_ID,STATION_NAME FROM station_info where STATION_ID=%s" % (station_id)
    # print(fetch_station)
    cur.execute(fetch_station)
    ctx = cur.fetchone()
    print(ctx)
    if ctx:
        _shift_emp_prod_pay(cur, ctx, shift_no, shift_date)
        _shift_prod_noz_order(cur, ctx, shift_no, shift_date)
        conn.commit()
        return _create_excel(cur, ctx, shift_no)
