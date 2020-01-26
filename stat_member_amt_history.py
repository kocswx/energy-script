# coding:utf-8
from openpyxl import Workbook
from config.config_db_server import *
from _openpyxl_constant import *
from _util import *
import datetime

sheet_title = '会员余额汇总'
cell_name = ['会员ID', '会员姓名', '账户类型', '状态', '所属企业名称', '卡号', '开始余额', '充值金额', '消费金额',
             '退款金额', '转出金额', '转入金额', '账户余额', '所属站点ID', '所属油站名称']

cur = conn.cursor()


class AmountType:
    start_amount = 0
    charge_amount = 0
    consume_amount = 0
    refund_amount = 0
    trans_out_amount = 0
    trans_in_amount = 0
    end_amount = 0


def _get_member_count(group_id, station_id, end_time):
    sql = "SELECT COUNT(1) from eng_crm.member_info as mi " \
          "INNER JOIN eng_crm.account as ma on mi.MEMBER_ID=ma.ACCOUNT_ID " \
          "WHERE mi.GROUP_ID=%s AND mi.CREATED_TIME<'%s' " % (group_id, end_time)
    if station_id:
        sql = sql + 'and mi.OPT_STATION_ID=%s' % station_id
    print(sql)
    cur.execute(sql)
    rs = cur.fetchone()
    return rs[0]


def _get_member_record(group_id, end_time, start_index, ps, station_id):
    sql = "SELECT mi.MEMBER_ID,mi.MEMBER_NAME,mi.MEMBER_TYPE,mi.`STATUS`,mi.COMPANY_NAME,mi.CARD_NO,mi.OPT_STATION_ID," \
          "mi.OPT_STATION_NAME,ma.BALANCE_AMT from eng_crm.member_info as mi INNER JOIN eng_crm.account as ma " \
          "on mi.MEMBER_ID=ma.ACCOUNT_ID  WHERE GROUP_ID='%s' AND CREATED_TIME<'%s' " % (group_id, end_time)
    if station_id:
        sql = sql + 'and mi.OPT_STATION_ID=%s' % station_id
    sql += " order by mi.MEMBER_TYPE,mi.MEMBER_ID limit %d,%d" % (start_index, ps)
    cur.execute(sql)
    return cur.fetchall()


def _write_excel(group_id, begin_time, end_time, station_id, emp_name, save_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
    row_index = 1

    sheet.merge_cells("A1:%s1" % column_letter[len(cell_name) - 1])
    set_cell(sheet, row_index, 1, sheet.title, font_rpt_title, align_center, border)
    row_index += 1

    set_cell(sheet, row_index, 1, ('统计时间：%s 到 %s' % (begin_time, end_time)), font_rpt_normal, align_left, border)
    sheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=5)
    set_cell(sheet, row_index, 6, ('提交员工：%s' % emp_name), font_rpt_normal, align_center, border)
    sheet.merge_cells(start_row=row_index, start_column=6, end_row=row_index, end_column=10)
    set_cell(sheet, row_index, 11, ('生成时间：%s' % get_now_full()), font_rpt_normal, align_right, border)
    sheet.merge_cells(start_row=row_index, start_column=11, end_row=row_index, end_column=15)

    row_index += 1
    for i in range(len(cell_name)):
        set_cell(sheet, row_index, i + 1, cell_name[i], font_rpt_cell_header, align_center, border)
        sheet.column_dimensions[column_letter[i]].width = 12
    row_index += 1

    member_count = _get_member_count(group_id, station_id, end_time)
    print("member_count:", member_count)
    member_list = _get_member_record(group_id, end_time, 0, member_count, station_id)
    # print(member_list)

    for item in member_list:
        member_id = item[0]
        print(member_id)
        balance_amt = item[8]
        amount_type = _get_member_all_type_amount(member_id, balance_amt, begin_time, end_time)
        # print(amount_type.__dict__)
        set_cell(sheet, row_index, 1, member_id, font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 2, item[1], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 3, MEMBER_TYPE_DICT[item[2]], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 4, MEMBER_STATUS_DICT[item[3]], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 5, item[4], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 6, item[5], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 7, amount_type.start_amount, font_rpt_normal, align_center, border)  # 开始余额
        set_cell(sheet, row_index, 8, amount_type.charge_amount, font_rpt_normal, align_center, border)  # 充值金额
        set_cell(sheet, row_index, 9, amount_type.consume_amount, font_rpt_normal, align_center, border)  # 消费金额
        set_cell(sheet, row_index, 10, amount_type.refund_amount, font_rpt_normal, align_center, border)  # 退款金额
        set_cell(sheet, row_index, 11, amount_type.trans_out_amount, font_rpt_normal, align_center, border)  # 转出金额
        set_cell(sheet, row_index, 12, amount_type.trans_in_amount, font_rpt_normal, align_center, border)  # 转入金额
        set_cell(sheet, row_index, 13, amount_type.end_amount, font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 14, item[6], font_rpt_normal, align_center, border)
        set_cell(sheet, row_index, 15, item[7], font_rpt_normal, align_center, border)
        row_index += 1

    wb.save(save_path)


def _get_member_all_type_amount(member_id, balance_amt, begin_time, end_time):
    amount_type = AmountType()
    amount_type.start_amount = _get_member_start_amount(member_id, begin_time, balance_amt)
    amount_type.end_amount = _get_member_end_amount(member_id, end_time, balance_amt)
    sql = "SELECT AMT_TYPE,SUM(AMT) from eng_crm.history_amt WHERE MEMBER_ID='%s' " \
          "and created_time>='%s' AND created_time<'%s' GROUP BY AMT_TYPE;" % (member_id, begin_time, end_time)
    cur.execute(sql)
    rs = cur.fetchall()
    if rs:
        for item in rs:
            amt_type = item[0]
            amt_value = item[1]
            # print(member_id, amt_type, amt_value)
            if amt_type == 1:  # 加油
                amount_type.consume_amount += amt_value
                continue
            if amt_type == 2:  # 充值
                amount_type.charge_amount += amt_value
                continue
            if amt_type == 3:  # 退款
                amount_type.refund_amount += amt_value
                continue
            if amt_type == 4:  # 转出
                amount_type.trans_out_amount += amt_value
                continue
            if amt_type == 5:  # 转入
                amount_type.trans_in_amount += amt_value
                continue
    return amount_type


def _get_member_start_amount(member_id, begin_time, balance_amount):
    sql = "select BALANCE_AMT-AMT from eng_crm.history_amt where MEMBER_ID='%s' and CREATED_TIME>='%s'" \
          " order by CREATED_TIME limit 1" % (member_id, begin_time)
    cur.execute(sql)
    rs = cur.fetchall()
    if rs is None or len(rs) == 0:
        sql = "select BALANCE_AMT, AMT from eng_crm.history_amt where MEMBER_ID='%s' and CREATED_TIME<'%s'" \
              " order by CREATED_TIME desc limit 1" % (member_id, begin_time)
        cur.execute(sql)
        rs2 = cur.fetchall()
        if rs2 is None or len(rs2) == 0:
            amount = balance_amount
        else:
            amount = rs2[0][0]
    else:
        amount = rs[0][0]
    return amount


def _get_member_end_amount(member_id, end_time, balance_amount):
    sql = "select BALANCE_AMT from eng_crm.history_amt where MEMBER_ID='%s' and created_time<='%s' " \
          "order by created_time desc limit 1" % (member_id, end_time)
    cur.execute(sql)
    rs = cur.fetchall()
    if rs is None or len(rs) == 0:
        sql = "select BALANCE_AMT-AMT from eng_crm.history_amt where MEMBER_ID='%s' and created_time>'%s' " \
              "order by created_time limit 1" % (member_id, end_time)
        cur.execute(sql)
        rs2 = cur.fetchall()
        if rs2 is None or len(rs2) == 0:
            amount = balance_amount
        else:
            amount = rs2[0][0]
    else:
        amount = rs[0][0]
    return amount


if __name__ == '__main__':
    print("member_amt_history_excel start: %s" % (datetime.datetime.now()))
    cur.execute("select ID,GROUP_ID,BEGIN_TIME,END_TIME,STATION_ID,EMP_NAME from eng_report.stat_job "
                "where REPORT_TYPE='memAmt' and CALC_REPORT=0 limit 1")
    job = cur.fetchone()
    if job:
        print(job)
        # /opt/eng_static/report/10003/memAmt-25.xlsx
        now_year_month = get_now_year_month()
        save_path = "%s/%s/%s/memAmt-%d.xlsx" % (save_report_folder, job[1], now_year_month, job[0])
        return_path = "%s/%s/%s/memAmt-%d.xlsx" % (return_report_prefix, job[1], now_year_month, job[0])
        print(save_path)
        prepare_path(save_path)
        station_id = job[4]
        if station_id is None or station_id <= 0:
            station_id = None
        _write_excel(job[1], job[2], job[3], station_id, job[5], save_path)
        cur.execute("update eng_report.stat_job set CALC_REPORT=1,REPORT_PATH='%s' where ID=%d" % (return_path, job[0]))
        conn.commit()
    print("member_amt_history_excel end: %s" % (datetime.datetime.now()))
    cur.close()
    conn.close()
